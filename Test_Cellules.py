#!/usr/bin/env python
# coding: utf-8

# In[ ]:


from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, NamedStyle
from openpyxl.worksheet.datavalidation import DataValidation


# In[ ]:


#Lire le fichier profil
profil = load_workbook('profil.xlsx')
#Vérouillage des cellules
#Pour le verouillage et la protection des cellules, d'après ce que j'ai trouvé ca marche comme suite: 
#il faut verouiller toute la feuille et puis déverouiller les cellules qu'on veut laisser accès à modifier
ws = profil.sheetnames
#Définir la feuille ciblée
ws[0] = profil.active
#Verouiller la sheet
ws[0].protection.sheet = True
#Indiquer la cellule à déverouiller
c = profil.active['A3']
#Déverouiller la cellule
c.protection = Protection(locked=False)
#Saving 
profil.save( "profil.xlsx" )


# In[ ]:


################################################################################################styles
#### Concernant le style, c'est possible pour les fichiers excel créés donc on pourra caractériser notre workbook comme voulu 
#### lorsqu'on créera le fichier profil et non lorsqu'on le chargera
#### cette section donc à insérer dans notre code source plus précisemment juste après la commande de création du fichier profil
for x in range(len(profil.sheetnames)):
    profil.active = x 
    row = profil.active.row_dimensions[1]
    row.font = Font(bold=True, size=13)
for x in range(len(profil.sheetnames)):
    profil.active = x 
    row = profil.active.row_dimensions[2]
    row.font = Font(bold=True, size=13)
for x in range(1,len(profil.sheetnames)):
    profil.active = x
    for row in profil.active.iter_rows(min_row=3, min_col=1, max_col=1, max_row=len(etl.header(etl.fromcsv( repertoireSource+filenames[x+1], delimiter = delimiter, quotechar = quotechar, quoting = quoting )))):
        for cell in row:
            cell.fill = PatternFill(fill_type='solid', fgColor='E0E0E0')
            bd = Side(style='thin', color="000000")
            cell.border = Border(left=None, top=bd, right=None, bottom=bd)
for x in range(1,len(profil.sheetnames)):
    profil.active = x
    for row in profil.active.iter_rows(min_row=3, min_col=3, max_col=4, max_row=len(etl.header(etl.fromcsv( repertoireSource+filenames[x+1], delimiter = delimiter, quotechar = quotechar, quoting = quoting )))):
        for cell in row:
            cell.fill = PatternFill(fill_type='solid', fgColor='E0E0E0')
            bd = Side(style='thin', color="000000")
            cell.border = Border(left=None, top=bd, right=None, bottom=bd)
for x in range(1,len(profil.sheetnames)):
    profil.active = x
    for row in profil.active.iter_rows(min_row=3, min_col=2, max_col=2, max_row=len(etl.header(etl.fromcsv( repertoireSource+filenames[x+1], delimiter = delimiter, quotechar = quotechar, quoting = quoting )))):
        for cell in row:
            cell.fill = PatternFill(fill_type='solid', fgColor='99B5D5')
            bd = Side(style='thin', color="000000")
            cell.border = Border(left=None, top=bd, right=None, bottom=bd)
for x in range(1,len(profil.sheetnames)):
    profil.active = x
    for row in profil.active.iter_rows(min_row=3, min_col=5, max_col=10, max_row=len(etl.header(etl.fromcsv( repertoireSource+filenames[x+1], delimiter = delimiter, quotechar = quotechar, quoting = quoting )))):
        for cell in row:
            cell.fill = PatternFill(fill_type='solid', fgColor='99B5D5')
            bd = Side(style='thin', color="000000")
            cell.border = Border(left=None, top=bd, right=None , bottom=bd)


# In[ ]:


#DataValidation
wb = Workbook()
ws = wb.active
#On definie notre liste déroulante
dv = DataValidation(type="list", formula1='"a,b,c"', allow_blank=True)
#On choisit les cellules où on veut appliquer notre datavalidation
c1 = ws["A1"]
c2 = ws["B1"]
dv.add(c1)
dv.add(c2)
ws.add_data_validation(dv) 
#Message d'invite
dv.prompt = 'Veuillez choisir parmi les valeurs de la liste'
dv.promptTitle = 'Liste de Selection'
#Message d'erreur
dv.error ='Valeur non valide'
dv.errorTitle = 'Entrée invalide!!!'
#saving
wb.save("test.xlsx")

