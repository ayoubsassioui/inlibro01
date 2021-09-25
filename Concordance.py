#!/usr/bin/env python
# coding: utf-8

# In[1]:


import petl as etl
import numpy as np
import csv
from os import path
from openpyxl import Workbook, load_workbook
from nltk import WordNetLemmatizer


# In[2]:


filename = input('Saisir le nom du fichier  :')     #Lexpl:          PLES_CI_DT_SUBSCRIBER
C = input('Saisir le nom de la colonne  :')         #Lexpl:          address_2


# In[9]:


#On va traiter la colonne address2 du fichier subscriber comme exemple
e = etl.fromcsv(filename , delimiter = '|', quotechar = "'", quoting = csv.QUOTE_ALL , encoding="utf8")


# In[10]:


l = list(e[C])
#Filtrer la liste des éléments vides
l1 = list(filter(None, l))
#rendre tous les elements de la liste en miniscule chose qui va réduire les occurences
l1 = [x.lower() for x in l1]
#Ordonner la liste en ordre alphabétique
l2 = sorted(l1)
# Extraire les valeurs distinctes avec le nombre d'occurence de chaque valeur
(unique, counts) = np.unique(l2, return_counts=True)
l3 = np.asarray((unique, counts)).T


# In[11]:


l3


# In[12]:


wl = WordNetLemmatizer()
w = []
for i in range(len(l3)):
    w.append(wl.lemmatize(l3[i][0]))


# In[13]:


#Charger le fichier profil et activer la Worksheet Correspondance
profil = load_workbook('profil.xlsx')
ws = profil.worksheets
profil.active = ws[2]

#Demander à l'usager de saisir le label
label = input('Saisir le label :')
profil.active['A1'] = label

#Remplir la colonne par les valeurs distinctes detectées
x = 0
for row in ws[2].iter_rows(min_row=2, min_col=1, max_col=1, max_row=1+len(l3)):
    for cell in row:
        cell.value = l3[x][0]
        x+= 1
        
#Remplir la colonne par le nombre d'occurence de ces valeurs
x = 0
for row in ws[2].iter_rows(min_row=2, min_col=2, max_col=2, max_row=1+len(l3)):
    for cell in row:
        cell.value = l3[x][1]
        x+= 1
#Remplir la colonne par la suggestion de la normalization
x = 0
for row in ws[2].iter_rows(min_row=2, min_col=3, max_col=3, max_row=1+len(l3)):
    for cell in row:
        cell.value = w[x]
        x+= 1

profil.save('Profil.xlsx')


# In[ ]:




