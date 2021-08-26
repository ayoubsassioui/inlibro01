#!/usr/bin/env python
# coding: utf-8

# In[ ]:


import os
import sys
import getopt
import petl as etl
if sys.platform == 'linux':
    import magic
import csv
from os import path


# In[ ]:


# Pour la génération d'un fichier Excel
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, NamedStyle


# In[ ]:


## Fonction qui permet de lire les fichiers selon la racine et la version
def generateProfilFromFiles( racine, version ):
    repertoireSource = racine + '/source/' + version + '/'
    assert( path.isdir( repertoireSource )), "Le chemin '" + repertoireSource + "' doit être un répertoire"

    filenames = os.listdir( repertoireSource )
    filesInfo = {}

    delimiter = '|'
    quotechar = "'"
    quoting = csv.QUOTE_ALL

    for filename in sorted( filenames ):

        # Ignorer les fichiers temporaires
        if filename.startswith("."):
            continue

        validData = False
        
        while not validData:
            choice = None
            print ("\n" + '=' * (len(filename) + 4) )
            print ("= " + filename + " =")
            print ( '=' * (len(filename) + 4) + "\n")
            filePath = repertoireSource+filename

            # Get file type
            ## TODO Ayoub : Remplacer la ligne suivante pour demander le type de fichier.  Mettre une condition 
            filetype = None
            if sys.platform == 'linux':
                filetype = magic.from_file( filePath )
            else:
                f = input(str("Saisir le type du fichier: [M]ARC21, [C]sv, [T]ext, [X]ml :  M , C , T , X ? ")) 
                while f not in ('M' , 'C' , 'T' , 'X'):
                    print("Type inconnu, veuillez resaisir le type du fichier!!")
                if f == 'M':
                    filetype = 'MARC21'
                elif f == 'T':
                    filetype = 'text'
                elif f == 'C':
                    filetype = 'CSV'
                elif f == 'X':
                    filetype = 'xml'
                    
                filesInfo[filename] = {}
                filesInfo[filename]['type'] = filetype
            
            if "MARC21" in filetype:
                print("Fichier MARC21: ignoré")
                validData = True
                filesInfo[filename]['specs'] = 'Ne pas traiter'
                choice = 'i' # on ignore pour l'instant

            if "text" in filetype:
                print('Fichier texte')
                table = etl.fromcsv( filePath, delimiter = delimiter, quotechar = quotechar, quoting = quoting )
                headers = etl.header( table )
                print( table.head( 3 ) )

                print("Séparateur de champs : " + ("Aucun" if delimiter is None else delimiter))
                print("Délimiteur de chaîne : " + ("Aucun" if quotechar is None else quotechar))

                print ("Nombre de colonnes du fichier : " + str(len(headers)))
                userInput = None
                while userInput not in ('o', 'n'):
                    userInput = input("Est-ce exacte (O/N)? ").lower()

                if userInput == 'o':
                    validData = True
                    filesInfo[filename]['type'] = 'CSV'
                    filesInfo[filename]['specs'] = "Séparateur={},Délimiteur={}".format(
                                                                                    delimiter if delimiter else "Aucun", 
                                                                                    quotechar if quotechar else "Aucun")
                    filesInfo[filename]['count'] = etl.nrows(table)
                    filesInfo[filename]['columns'] = {}
                    for column in etl.header(table):
                        filesInfo[filename]['columns'][column] = {}
                        counter = etl.valuecounter(table, column)
                        filesInfo[filename]['columns'][column]['distinct'] = len( counter )
                        filesInfo[filename]['columns'][column]['empty'] = 0
                        for emptyValue in (0, '', None, '""'):
                            filesInfo[filename]['columns'][column]['empty'] += counter[emptyValue]

                    print( table )
                elif userInput == 'n':
                    choice, delimiter, quotechar = askForCSVSpecs(delimiter=delimiter, quotechar=quotechar)
                    validData = True if choice in ('i', 'n') else False
                    quoting = csv.QUOTE_NONE if quotechar is None else csv.QUOTE_MINIMAL
                else:
                    print("Choix invalide")
                    continue

            if choice == 'i':
                filesInfo[filename]['specs'] : 'Ignoré'
            elif choice == 'n':
                filesInfo[filename]['specs'] : 'Ne pas traiter'
    return filesInfo


# In[ ]:


def askForCSVSpecs( delimiter, quotechar):
    choice = None
    while choice not in ('c', 'i', 'n'):
        question = "Que voulez-vous modifier ?\n[S]éparateur : {} \n[D]élimiteur de champs : {}\n[I]gnorer ce fichier pour l'instant\n[N]e pas traiter ce fichier\n[C]onfirmer\n(S/D/I/N/C) ? ".format(("'"+delimiter+"'" if delimiter else 'Aucun'), (quotechar if quotechar else 'Aucun'))
        choice = input(question).lower()
        if choice == 's':
            delimiter = input("Quel séparateur désirez-vous utiliser ? ")
        elif choice == 'd':
            quotechar = input("Quel délimiteur désirez-vous utiliser ? ")
    return choice, delimiter, quotechar


# In[ ]:


########
# Read command line parameters
########
argv = sys.argv[1:]

options, args = getopt.getopt(argv,'r:v:s:',["racine=","version=","sigb="])

racine = None
version = None
sigb = None

for option, value in options:
    print( '['+option+':'+value+']')
    if option in ('-r', '--racine'):
        racine = value
    elif option in ('-v', '--version'):
        version = value
    elif option in ('-s', '--sigb'):
        sigb = value

########
# Read required parameters not specified on commandline
########
if racine is None:
    erreur = "--racine non spécifié"
    while erreur :
        racine = input(erreur + "\nRépertoire racine à utiliser ? ")
        erreur = ""
        if not path.isdir(racine + "/source"):
            erreur = "Répertoire invalide : "+ racine + "/source n'existe pas"

if version not in ('test', 'prod'):
    while True:
        valeur = input(str('Version à utiliser [T]est ou [P]rod\n(T/P)? ')).lower();

        if valeur not in ('t', 'p'):
            print ("Choix invalide")
        else:
            version = "test" if valeur == 't' else "prod"
            break

if sigb is None:
    while True:
        sigb = input(str('Quel SIGB ? '))
        
        valeur = input("Vous avez saisie : " + sigb + "\nEst-ce exact ? (O/N) ").lower()
        if valeur == 'o':
            break

# Read the file content if the file existe
## TODO Ayoub : Vérifier si le fichier profil.xslx existe, si oui, le lire, sinon, on le génère à partir des fichiers
##              Dans les deux cas, le contenu de filesInfo devrait être le même... pour l'instant

if path.exists("profil.xslx") == True :
    profil = load_workbook('profil.xlsx')
    filesInfo = generateProfilFromFiles( racine, version )
    print(filesInfo)
    
    
else:
    filesInfo = generateProfilFromFiles( racine, version )
    print(filesInfo)
    # Create the Workbook
    profil = Workbook()
    fichiers = profil.active
    fichiers.title = ("Fichiers")

    emptyLine = []
    fournisseur = [ "Source", sigb ]
    enteteFichier = [ "Fichier", "Type", "Info", "Quantité de données", "Données", "Clé du fichier", "Fichier à fusionner", "Clé à utiliser"]
    fichiers.append( fournisseur )
    fichiers.append( emptyLine )
    fichiers.append( enteteFichier )

    for filename in filesInfo.keys():
        skip = 'count' not in filesInfo[filename]
        fichiers.append([ filename, filesInfo[filename]['type'], filesInfo[filename]['specs'], "" if skip else filesInfo[filename]['count'] ])

        # Ajout des feuille pour chaque fichier
        worksheet = profil.create_sheet(filename)
        worksheet['H1'] = "Filtre"
        worksheet.append(['Champs', 'Clé', 'Nombre de valeurs distinctes', 'Nombre de valeurs vides', 'Type de données', 'Champs Koha', 'Correspondance', 'Opérateur', 'Valeur 1', 'Valeur 2' ])
        if 'columns' in filesInfo[filename]:
            for column in filesInfo[filename]['columns']:
                worksheet.append([column, "", filesInfo[filename]['columns'][column]['distinct'], filesInfo[filename]['columns'][column]['empty'] ])

    profil.save( racine + "/profil.xlsx" )

